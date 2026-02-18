#!/usr/bin/env python3
"""
Gera planilha Excel com abas: Transações, Categorias, Instituições.
Validação: na aba Transações, os campos Instituição e Categoria são listas
suspensas que usam os nomes cadastrados nas outras abas.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


def criar_planilha(caminho_saida: str = "planilha_transacoes.xlsx") -> None:
    wb = Workbook()

    # ---- ABA INSTITUIÇÕES (criar primeiro para as referências) ----
    ws_inst = wb.active
    ws_inst.title = "Instituicoes"
    headers_inst = ["Nome", "Tipo"]
    for col, h in enumerate(headers_inst, start=1):
        cell = ws_inst.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    # Validação para Tipo na própria aba (opcional)
    dv_tipo_inst = DataValidation(
        type="list",
        formula1='"Banco,Vale"',
        allow_blank=True,
    )
    dv_tipo_inst.error = "Escolha um tipo da lista"
    ws_inst.add_data_validation(dv_tipo_inst)
    dv_tipo_inst.add("B2:B500")

    # ---- ABA CATEGORIAS ----
    ws_cat = wb.create_sheet("Categorias", 1)
    headers_cat = ["Nome", "Tipo"]
    for col, h in enumerate(headers_cat, start=1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    dv_tipo_cat = DataValidation(
        type="list",
        formula1='"Gasto,Receita,Global"',
        allow_blank=False,
    )
    ws_cat.add_data_validation(dv_tipo_cat)
    dv_tipo_cat.add("B2:B500")

    # ---- ABA TRANSAÇÕES ----
    ws_trans = wb.create_sheet("Transacoes", 2)
    headers_trans = [
        "Data",
        "Tipo",
        "Valor",
        "Descricao",
        "Instituicao",
        "Categoria",
        "Parcelado",
        "Recorrencia",
        "Fim da recorrencia",
    ]
    for col, h in enumerate(headers_trans, start=1):
        cell = ws_trans.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Referência direta à aba (funciona no Excel; no Google Planilhas veja LEIA-ME.txt)
    # Lista suspensa Instituição = nomes da aba Instituicoes (coluna A)
    dv_inst = DataValidation(
        type="list",
        formula1="=Instituicoes!$A$2:$A$500",
        allow_blank=True,
    )
    dv_inst.error = "Selecione uma instituição cadastrada na aba Instituicoes."
    ws_trans.add_data_validation(dv_inst)
    dv_inst.add("E2:E5000")  # coluna Instituicao

    # Lista suspensa Categoria = nomes da aba Categorias (coluna A)
    dv_cat = DataValidation(
        type="list",
        formula1="=Categorias!$A$2:$A$500",
        allow_blank=True,
    )
    dv_cat.error = "Selecione uma categoria cadastrada na aba Categorias."
    ws_trans.add_data_validation(dv_cat)
    dv_cat.add("F2:F5000")  # coluna Categoria

    # Lista suspensa Tipo (Gasto/Receita)
    dv_tipo_trans = DataValidation(
        type="list",
        formula1='"Gasto,Receita"',
        allow_blank=False,
    )
    ws_trans.add_data_validation(dv_tipo_trans)
    dv_tipo_trans.add("B2:B5000")

    # Lista suspensa Parcelado
    dv_parcelado = DataValidation(
        type="list",
        formula1='"Sim,Não"',
        allow_blank=True,
    )
    ws_trans.add_data_validation(dv_parcelado)
    dv_parcelado.add("G2:G5000")

    # Lista suspensa Recorrência
    dv_rec = DataValidation(
        type="list",
        formula1='"Diário,Semanal,Mensal,Anual"',
        allow_blank=True,
    )
    ws_trans.add_data_validation(dv_rec)
    dv_rec.add("H2:H5000")

    # Largura das colunas (Transações)
    for col in range(1, len(headers_trans) + 1):
        ws_trans.column_dimensions[get_column_letter(col)].width = 14
    ws_trans.column_dimensions["D"].width = 30  # Descrição
    ws_trans.column_dimensions["E"].width = 18
    ws_trans.column_dimensions["F"].width = 18

    wb.save(caminho_saida)
    print(f"Planilha salva em: {caminho_saida}")
    print("Abas: Instituicoes, Categorias, Transacoes.")
    print("Preencha primeiro Instituicoes e Categorias; em Transacoes use as listas suspensas.")


if __name__ == "__main__":
    criar_planilha()
