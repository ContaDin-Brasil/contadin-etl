#!/usr/bin/env python3
"""
Gera planilha Excel com abas: Transações, Categorias, Instituições,
JÁ PREENCHIDA com dados de exemplo (instituições, categorias e 15+ transações).
"""
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# Dados de exemplo
INSTITUICOES = [
    ("Nubank", "Banco"),
    ("Itaú", "Banco"),
    ("Vale Refeição", "Vale"),
    ("Vale Transporte", "Vale"),
]

CATEGORIAS = [
    ("Alimentação", "Gasto"),
    ("Transporte", "Gasto"),
    ("Moradia", "Gasto"),
    ("Lazer", "Gasto"),
    ("Salário", "Receita"),
    ("Freelance", "Receita"),
    ("Outros gastos", "Global"),
]

# 15+ transações: (data, tipo, valor, descricao, instituicao, categoria, parcelado, recorrencia, fim_recorrencia)
# instituicao e categoria devem ser nomes que existem em INSTITUICOES e CATEGORIAS
HOJE = datetime.now().date()
TRANSACOES = [
    (HOJE - timedelta(days=2), "Gasto", 45.90, "Supermercado", "Nubank", "Alimentação", "Não", "", None),
    (HOJE - timedelta(days=1), "Gasto", 12.00, "Ônibus", "Nubank", "Transporte", "Não", "Mensal", None),
    (HOJE, "Receita", 3500.00, "Salário mensal", "Itaú", "Salário", "Não", "Mensal", None),
    (HOJE - timedelta(days=5), "Gasto", 28.50, "Almoço", "Vale Refeição", "Alimentação", "Não", "", None),
    (HOJE - timedelta(days=4), "Gasto", 1200.00, "Aluguel", "Itaú", "Moradia", "Não", "Mensal", None),
    (HOJE - timedelta(days=6), "Gasto", 35.00, "Cinema", "Nubank", "Lazer", "Não", "", None),
    (HOJE - timedelta(days=3), "Receita", 500.00, "Projeto extra", "Nubank", "Freelance", "Não", "", None),
    (HOJE - timedelta(days=7), "Gasto", 89.90, "Farmácia", "Nubank", "Outros gastos", "Não", "", None),
    (HOJE - timedelta(days=8), "Gasto", 15.00, "Uber", "Nubank", "Transporte", "Não", "", None),
    (HOJE - timedelta(days=9), "Gasto", 22.00, "Lanche", "Vale Refeição", "Alimentação", "Não", "", None),
    (HOJE - timedelta(days=10), "Gasto", 350.00, "Conta de luz", "Itaú", "Moradia", "Não", "Mensal", None),
    (HOJE - timedelta(days=11), "Gasto", 199.90, "Assinatura streaming", "Nubank", "Lazer", "Sim", "Mensal", None),
    (HOJE - timedelta(days=12), "Gasto", 8.50, "Metrô", "Vale Transporte", "Transporte", "Não", "", None),
    (HOJE - timedelta(days=13), "Gasto", 55.00, "Restaurante", "Vale Refeição", "Alimentação", "Não", "", None),
    (HOJE - timedelta(days=14), "Receita", 3500.00, "Salário", "Itaú", "Salário", "Não", "Mensal", None),
    (HOJE - timedelta(days=15), "Gasto", 42.00, "Posto gasolina", "Nubank", "Transporte", "Não", "", None),
    (HOJE - timedelta(days=1), "Gasto", 18.90, "iFood", "Nubank", "Alimentação", "Não", "", None),
]


def criar_planilha_exemplo(caminho_saida: str = "planilha_transacoes_exemplo.xlsx") -> None:
    wb = Workbook()

    # ---- ABA INSTITUIÇÕES ----
    ws_inst = wb.active
    ws_inst.title = "Instituicoes"
    headers_inst = ["Nome", "Tipo"]
    for col, h in enumerate(headers_inst, start=1):
        cell = ws_inst.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    for row, (nome, tipo) in enumerate(INSTITUICOES, start=2):
        ws_inst.cell(row=row, column=1, value=nome)
        ws_inst.cell(row=row, column=2, value=tipo)
    dv_tipo_inst = DataValidation(
        type="list",
        formula1='"Banco,Vale"',
        allow_blank=True,
    )
    dv_tipo_inst.error = "Escolha um tipo da lista"
    ws_inst.add_data_validation(dv_tipo_inst)
    dv_tipo_inst.add("B2:B500")

    # ---- ABA CATEGORIAS ----
    ws_cat = wb.create_sheet("Categorias", 1)
    headers_cat = ["Nome", "Tipo"]
    for col, h in enumerate(headers_cat, start=1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    for row, (nome, tipo) in enumerate(CATEGORIAS, start=2):
        ws_cat.cell(row=row, column=1, value=nome)
        ws_cat.cell(row=row, column=2, value=tipo)
    dv_tipo_cat = DataValidation(
        type="list",
        formula1='"Gasto,Receita,Global"',
        allow_blank=False,
    )
    ws_cat.add_data_validation(dv_tipo_cat)
    dv_tipo_cat.add("B2:B500")

    # ---- ABA TRANSAÇÕES ----
    ws_trans = wb.create_sheet("Transacoes", 2)
    headers_trans = [
        "Data",
        "Tipo",
        "Valor",
        "Descricao",
        "Instituicao",
        "Categoria",
        "Parcelado",
        "Recorrencia",
        "Fim da recorrencia",
    ]
    for col, h in enumerate(headers_trans, start=1):
        cell = ws_trans.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for row, t in enumerate(TRANSACOES, start=2):
        data, tipo, valor, desc, inst, cat, parcelado, rec, fim_rec = t
        ws_trans.cell(row=row, column=1, value=data)
        ws_trans.cell(row=row, column=2, value=tipo)
        ws_trans.cell(row=row, column=3, value=valor)
        ws_trans.cell(row=row, column=4, value=desc)
        ws_trans.cell(row=row, column=5, value=inst)
        ws_trans.cell(row=row, column=6, value=cat)
        ws_trans.cell(row=row, column=7, value=parcelado)
        ws_trans.cell(row=row, column=8, value=rec)
        if fim_rec:
            ws_trans.cell(row=row, column=9, value=fim_rec)
        else:
            ws_trans.cell(row=row, column=9, value="")

    # Validações (listas suspensas) para novas linhas
    dv_inst = DataValidation(
        type="list",
        formula1="=Instituicoes!$A$2:$A$500",
        allow_blank=True,
    )
    dv_inst.error = "Selecione uma instituição cadastrada na aba Instituicoes."
    ws_trans.add_data_validation(dv_inst)
    dv_inst.add("E2:E5000")

    dv_cat = DataValidation(
        type="list",
        formula1="=Categorias!$A$2:$A$500",
        allow_blank=True,
    )
    dv_cat.error = "Selecione uma categoria cadastrada na aba Categorias."
    ws_trans.add_data_validation(dv_cat)
    dv_cat.add("F2:F5000")

    dv_tipo_trans = DataValidation(
        type="list",
        formula1='"Gasto,Receita"',
        allow_blank=False,
    )
    ws_trans.add_data_validation(dv_tipo_trans)
    dv_tipo_trans.add("B2:B5000")

    dv_parcelado = DataValidation(
        type="list",
        formula1='"Sim,Não"',
        allow_blank=True,
    )
    ws_trans.add_data_validation(dv_parcelado)
    dv_parcelado.add("G2:G5000")

    dv_rec = DataValidation(
        type="list",
        formula1='"Diário,Semanal,Mensal,Anual"',
        allow_blank=True,
    )
    ws_trans.add_data_validation(dv_rec)
    dv_rec.add("H2:H5000")

    for col in range(1, len(headers_trans) + 1):
        ws_trans.column_dimensions[get_column_letter(col)].width = 14
    ws_trans.column_dimensions["D"].width = 30
    ws_trans.column_dimensions["E"].width = 18
    ws_trans.column_dimensions["F"].width = 18

    wb.save(caminho_saida)
    print(f"Planilha de exemplo salva em: {caminho_saida}")
    print(f"  Instituicoes: {len(INSTITUICOES)} cadastradas.")
    print(f"  Categorias: {len(CATEGORIAS)} cadastradas.")
    print(f"  Transacoes: {len(TRANSACOES)} de exemplo.")


if __name__ == "__main__":
    criar_planilha_exemplo()
